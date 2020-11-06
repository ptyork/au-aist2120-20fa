import openpyxl as xl

# workbook (xlsx file)
#   worksheet (1..n)
#     rows (1..n)
#       cell (1..n)

wb = xl.load_workbook('test.xlsx')   # returns an open workbook object

print(wb.worksheets)      # list of all worksheets
print(wb.sheetnames)      # list of the NAMES of all worksheets
print(wb.active)          # the currently active worksheet
print(wb.worksheets[0])   # access to a worksheet by position
print(wb['Sheet1'])       # access to a worksheet by NAME

ws = wb.worksheets[0]

print('ROWS')
print(ws.rows)
print(list(ws.rows))
print('COLUMNS')
print(ws.columns)
print(list(ws.columns))

for row in ws.rows:    # each row is a tuple of cells
    for cell in row:   # each cell is a value
        print(cell.value)
        print(cell.col_idx)
        print(cell.column)
        print(cell.column_letter)
        print(cell.row)
    
print("MAX COL")
print(ws.max_column)
print("MAX ROW")
print(ws.max_row)

# openpyxl allows RANDOM ACCESS TO DATA

print(ws['A1'])         # a specific cell
print(ws['a1'].value)   # value of a specific cell
print(ws['1'])          # a specific row
print(ws['a'])          # a specific column
print(ws['B2:C5'])      # a specified range of cells


# iterate over a range
for row in ws['B2:C5']:
    for cell in row:
        print(f"{cell.column_letter}{cell.row}: {cell.value}")